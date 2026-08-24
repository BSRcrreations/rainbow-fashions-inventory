from __future__ import annotations

import csv
import hashlib
import io
import re
from collections import Counter
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.config import Settings, get_settings
from app.core.exceptions import bad_request, conflict, not_found
from app.models.brand import Brand
from app.models.category import Category
from app.models.enums import OpeningStockImportStatus, PricingType, StockMovementType
from app.models.opening_stock_import import OpeningStockImport, OpeningStockImportAudit, OpeningStockImportError, OpeningStockImportRow
from app.models.product import Product
from app.models.product_barcode import ProductBarcode
from app.models.product_inventory import ProductInventory
from app.models.product_variant import InventoryCostLot, ProductVariant
from app.models.stock_history import StockHistory
from app.models.subcategory import SubCategory
from app.models.user import User
from app.schemas.opening_stock_import import OPENING_STOCK_CONFIRMATION, OpeningStockImportConfirm, OpeningStockImportReport, OpeningStockImportReverse
from app.services.backup_status_service import BackupStatusService


REQUIRED_HEADERS = ("product_name", "category", "subcategory", "brand", "sku", "barcode", "quantity", "purchase_cost", "selling_price")
OPTIONAL_HEADERS = ("size", "color", "style_code", "mrp", "hsn_code", "gst_rate", "description", "unit", "warehouse")
ALL_HEADERS = set(REQUIRED_HEADERS + OPTIONAL_HEADERS)
SAFE_TEXT = re.compile(r"^[^\x00-\x1f]{1,500}$")


class OpeningStockImportService:
    """Strict, all-or-nothing opening-stock import with retained row-level evidence."""

    def __init__(self, db: Session, settings: Settings | None = None) -> None:
        self.db = db
        self.settings = settings or get_settings()

    async def upload_and_validate(self, upload: UploadFile, current_user: User, request_id: str | None) -> OpeningStockImport:
        store_id = self._store_id(current_user)
        filename = Path(upload.filename or "").name
        suffix = Path(filename).suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            raise bad_request("Upload a CSV or XLSX opening-stock file.", "UNSUPPORTED_OPENING_STOCK_FILE")
        content = await upload.read()
        if not content:
            raise bad_request("The opening-stock file is empty.", "EMPTY_OPENING_STOCK_FILE")
        if len(content) > self.settings.max_opening_stock_import_size_bytes:
            raise bad_request("The opening-stock file exceeds the configured size limit.", "OPENING_STOCK_FILE_TOO_LARGE")
        digest = hashlib.sha256(content).hexdigest()
        existing = self.db.query(OpeningStockImport).filter_by(store_id=store_id, file_sha256=digest).first()
        if existing:
            return existing

        records, file_errors = self._parse(filename, content)
        stored_filename = f"{uuid4()}{suffix}"
        self._write_evidence_file(stored_filename, content)
        batch = OpeningStockImport(
            store_id=store_id, uploaded_by=current_user.id, status=OpeningStockImportStatus.VALIDATING,
            original_filename=filename, stored_filename=stored_filename, content_type=upload.content_type,
            file_size_bytes=len(content), file_sha256=digest,
        )
        self.db.add(batch)
        self.db.flush()
        for code, message in file_errors:
            self._error(batch.id, None, None, None, code, message)
        seen = Counter()
        for number, raw in enumerate(records, start=2):
            normalized, errors = self._normalize_row(raw)
            for identity_field in ("barcode", "sku"):
                identity = normalized.get(identity_field, "")
                if identity:
                    seen[f"{identity_field}:{identity.casefold()}"] += 1
            row = OpeningStockImportRow(
                opening_stock_import_id=batch.id, row_number=number, raw_data=raw, normalized_data=normalized,
                validation_status="VALID" if not errors else "INVALID",
            )
            self.db.add(row)
            self.db.flush()
            for field, code, message in errors:
                self._error(batch.id, row.id, number, field, code, message)
        for row in self.db.query(OpeningStockImportRow).filter_by(opening_stock_import_id=batch.id).all():
            for identity_field in ("barcode", "sku"):
                key = row.normalized_data.get(identity_field, "")
                if key and seen[f"{identity_field}:{key.casefold()}"] > 1:
                    row.validation_status = "INVALID"
                    self._error(batch.id, row.id, row.row_number, identity_field, "DUPLICATE_ROW_IDENTITY", f"{identity_field.replace('_', ' ').title()} is duplicated in this file.")
        self._finish_validation(batch, current_user, request_id)
        self.db.commit()
        self.db.refresh(batch)
        return batch

    def list(self, current_user: User) -> list[OpeningStockImport]:
        return self.db.query(OpeningStockImport).filter_by(store_id=self._store_id(current_user)).order_by(OpeningStockImport.created_at.desc()).limit(100).all()

    def get(self, import_id: UUID, current_user: User) -> OpeningStockImport:
        batch = self.db.query(OpeningStockImport).filter_by(id=import_id, store_id=self._store_id(current_user)).first()
        if not batch:
            raise not_found("Opening stock import")
        return batch

    def detail_rows(self, import_id: UUID) -> tuple[list[OpeningStockImportRow], list[OpeningStockImportError]]:
        rows = self.db.query(OpeningStockImportRow).filter_by(opening_stock_import_id=import_id).order_by(OpeningStockImportRow.row_number).all()
        errors = self.db.query(OpeningStockImportError).filter_by(opening_stock_import_id=import_id).order_by(OpeningStockImportError.row_number).all()
        return rows, errors

    def confirm(self, import_id: UUID, payload: OpeningStockImportConfirm, current_user: User, request_id: str | None) -> OpeningStockImportReport:
        if payload.confirmation.strip() != OPENING_STOCK_CONFIRMATION:
            raise bad_request(f"Type {OPENING_STOCK_CONFIRMATION} to post this import.", "OPENING_STOCK_CONFIRMATION_REQUIRED")
        batch = self.get(import_id, current_user)
        if batch.status == OpeningStockImportStatus.COMPLETED:
            if batch.idempotency_key == payload.idempotency_key:
                return self._report(batch, already_completed=True)
            raise conflict("This import was already posted with a different idempotency key.", "OPENING_STOCK_ALREADY_POSTED")
        if batch.status != OpeningStockImportStatus.READY_TO_CONFIRM:
            raise conflict("Only a clean, backup-gated preview can be posted.", "OPENING_STOCK_NOT_READY")
        if not self._backup_gate()[0]:
            raise conflict("A current successful database backup is required before posting.", "OPENING_STOCK_BACKUP_REQUIRED")
        try:
            batch = self.db.query(OpeningStockImport).filter_by(id=batch.id).with_for_update().one()
            if batch.status == OpeningStockImportStatus.COMPLETED:
                if batch.idempotency_key == payload.idempotency_key:
                    return self._report(batch, already_completed=True)
                raise conflict("This import was already posted with a different idempotency key.", "OPENING_STOCK_ALREADY_POSTED")
            if batch.status != OpeningStockImportStatus.READY_TO_CONFIRM:
                raise conflict("The import changed while it was being confirmed. Refresh the preview.", "OPENING_STOCK_PREVIEW_STALE")
            batch.status = OpeningStockImportStatus.POSTING
            batch.idempotency_key = payload.idempotency_key
            rows = self.db.query(OpeningStockImportRow).filter_by(opening_stock_import_id=batch.id).order_by(OpeningStockImportRow.row_number).all()
            if len(rows) != batch.valid_row_count or batch.error_count:
                raise conflict("The preview is stale or contains errors. Upload and validate again.", "OPENING_STOCK_PREVIEW_STALE")
            post_result: Counter[str] = Counter()
            for row in rows:
                post_result.update(self._post_row(batch, row, current_user, request_id))
            batch.status = OpeningStockImportStatus.COMPLETED
            batch.confirmed_by = current_user.id
            batch.posted_at = datetime.now(timezone.utc)
            batch.validation_summary = {**batch.validation_summary, "post_result": dict(post_result)}
            self._audit(batch.id, "POSTED", current_user.id, request_id, {"row_count": len(rows), "idempotency_key": payload.idempotency_key})
            self.db.commit()
            self.db.refresh(batch)
            return self._report(batch)
        except Exception:
            self.db.rollback()
            raise

    def reverse(self, import_id: UUID, payload: OpeningStockImportReverse, current_user: User, request_id: str | None) -> OpeningStockImportReport:
        if payload.confirmation.strip() != "REVERSE OPENING STOCK":
            raise bad_request("Type REVERSE OPENING STOCK to reverse this import.", "OPENING_STOCK_REVERSAL_CONFIRMATION_REQUIRED")
        batch = self.get(import_id, current_user)
        if batch.status == OpeningStockImportStatus.REVERSED:
            return self._report(batch, already_completed=True)
        if batch.status != OpeningStockImportStatus.COMPLETED:
            raise conflict("Only a completed import can be reversed.", "OPENING_STOCK_NOT_COMPLETED")
        try:
            batch = self.db.query(OpeningStockImport).filter_by(id=batch.id).with_for_update().one()
            if batch.status == OpeningStockImportStatus.REVERSED:
                return self._report(batch, already_completed=True)
            if batch.status != OpeningStockImportStatus.COMPLETED:
                raise conflict("The import changed while it was being reversed. Refresh the preview.", "OPENING_STOCK_REVERSAL_STALE")
            rows = self.db.query(OpeningStockImportRow).filter_by(opening_stock_import_id=batch.id).order_by(OpeningStockImportRow.row_number).with_for_update().all()
            for row in rows:
                if not row.product_variant_id or not row.stock_history_id:
                    raise conflict("Import evidence is incomplete; reversal requires manual investigation.", "OPENING_STOCK_REVERSAL_EVIDENCE_MISSING")
                later = self.db.query(StockHistory.id).filter(
                    StockHistory.product_variant_id == row.product_variant_id,
                    StockHistory.created_at > batch.posted_at,
                    StockHistory.id != row.stock_history_id,
                ).first()
                if later:
                    raise conflict("This import has later stock activity and cannot be automatically reversed.", "OPENING_STOCK_REVERSAL_HAS_LATER_ACTIVITY")
                variant = self.db.query(ProductVariant).filter_by(id=row.product_variant_id).with_for_update().one()
                quantity = int(row.normalized_data["quantity"])
                if variant.current_stock < quantity:
                    raise conflict("Current stock is below the imported quantity; automatic reversal is unsafe.", "OPENING_STOCK_REVERSAL_NEGATIVE_STOCK")
                before = variant.current_stock
                variant.current_stock -= quantity
                product = self.db.query(Product).filter_by(id=variant.product_id).with_for_update().one()
                product.current_stock = max(0, product.current_stock - quantity)
                inventory = self.db.query(ProductInventory).filter_by(product_id=product.id, store_id=batch.store_id).with_for_update().first()
                if inventory:
                    inventory.current_stock = max(0, inventory.current_stock - quantity)
                self.db.add(StockHistory(product_id=product.id, product_variant_id=variant.id, store_id=batch.store_id, movement_type=StockMovementType.STOCK_RESET_OUT, qty=quantity, before_stock=before, after_stock=variant.current_stock, unit_cost=variant.average_cost, reference=f"Opening stock reversal {batch.id}", request_id=request_id, correction_of_id=row.stock_history_id, correction_reason="DUPLICATE_OPENING_STOCK", correction_notes=payload.reason, created_by=current_user.id))
            batch.status = OpeningStockImportStatus.REVERSED
            batch.reversed_by = current_user.id
            batch.reversed_at = datetime.now(timezone.utc)
            batch.reversal_reason = payload.reason
            self._audit(batch.id, "REVERSED", current_user.id, request_id, {"reason": payload.reason})
            self.db.commit()
            return self._report(batch)
        except Exception:
            self.db.rollback()
            raise

    def post_migration_opening_stock(
        self,
        *,
        product: Product,
        variant: ProductVariant,
        store_id: UUID,
        quantity: int,
        unit_cost: Decimal,
        current_user: User,
        reference: str,
        request_id: str,
    ) -> tuple[InventoryCostLot, StockHistory]:
        """Post one vetted migration quantity using the normal inventory evidence.

        The caller owns the surrounding transaction and must have performed the
        package, owner, environment, and idempotency checks.  This deliberately
        creates no purchase or TEST ledger data.
        """
        if quantity < 0:
            raise ValueError("Opening-stock quantity cannot be negative")
        if quantity == 0:
            raise ValueError("Zero opening-stock quantity does not create a movement")
        before = variant.current_stock
        variant.current_stock += quantity
        variant.last_purchase_cost = unit_cost
        variant.average_cost = ((variant.average_cost * before) + (unit_cost * quantity)) / (before + quantity) if before else unit_cost
        product.current_stock += quantity
        product.purchase_price = unit_cost
        inventory = self.db.query(ProductInventory).filter_by(product_id=product.id, store_id=store_id).with_for_update().first()
        if not inventory:
            inventory = ProductInventory(product_id=product.id, store_id=store_id, current_stock=0)
            self.db.add(inventory)
        inventory.current_stock += quantity
        lot = InventoryCostLot(store_id=store_id, product_variant_id=variant.id, received_quantity=quantity, remaining_quantity=quantity, unit_purchase_cost=unit_cost, effective_unit_cost=unit_cost, lot_reference=reference)
        self.db.add(lot)
        self.db.flush()
        movement = StockHistory(product_id=product.id, product_variant_id=variant.id, purchase_cost_lot_id=lot.id, unit_cost=unit_cost, store_id=store_id, movement_type=StockMovementType.OPENING_STOCK, qty=quantity, before_stock=before, after_stock=variant.current_stock, reference=reference, request_id=request_id, created_by=current_user.id)
        self.db.add(movement)
        self.db.flush()
        return lot, movement

    def _parse(self, filename: str, content: bytes) -> tuple[list[dict[str, str]], list[tuple[str, str]]]:
        try:
            if filename.lower().endswith(".csv"):
                decoded = content.decode("utf-8-sig")
                reader = csv.reader(io.StringIO(decoded))
                data = list(reader)
            else:
                workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
                if len(workbook.sheetnames) != 1:
                    return [], [("MULTIPLE_WORKSHEETS", "XLSX imports must contain exactly one worksheet.")]
                sheet = workbook.active
                data = []
                for values in sheet.iter_rows(values_only=False):
                    row: list[str] = []
                    for cell in values:
                        if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                            return [], [("FORMULA_NOT_ALLOWED", "Spreadsheet formulas are not allowed in opening-stock imports.")]
                        row.append("" if cell.value is None else str(cell.value))
                    data.append(row)
        except UnicodeDecodeError:
            return [], [("INVALID_ENCODING", "CSV must be UTF-8 encoded.")]
        except Exception:
            return [], [("UNREADABLE_FILE", "The uploaded spreadsheet could not be read safely.")]
        if not data:
            return [], [("EMPTY_FILE", "The uploaded file has no rows.")]
        headers = [str(value).strip().casefold().replace(" ", "_") for value in data[0]]
        if any(not value for value in headers):
            return [], [("EMPTY_HEADER", "Every column needs a header.")]
        duplicates = [name for name, count in Counter(headers).items() if count > 1]
        if duplicates:
            return [], [("DUPLICATE_HEADER", "Duplicate column headers are not allowed.")]
        missing = set(REQUIRED_HEADERS) - set(headers)
        unexpected = set(headers) - ALL_HEADERS
        if missing:
            return [], [("MISSING_REQUIRED_HEADER", f"Missing required columns: {', '.join(sorted(missing))}.")]
        if unexpected:
            return [], [("UNEXPECTED_HEADER", f"Unsupported columns: {', '.join(sorted(unexpected))}.")]
        body = data[1:]
        if len(body) > self.settings.max_opening_stock_import_rows:
            return [], [("ROW_LIMIT_EXCEEDED", f"Opening-stock imports are limited to {self.settings.max_opening_stock_import_rows:,} rows.")]
        if any(len(row) != len(headers) for row in body):
            return [], [("ROW_COLUMN_COUNT_MISMATCH", "Every data row must contain exactly the declared number of columns.")]
        return [dict(zip(headers, row)) for row in body if any(str(cell).strip() for cell in row)], []

    def _normalize_row(self, raw: dict[str, str]) -> tuple[dict[str, str], list[tuple[str | None, str, str]]]:
        data = {key: str(raw.get(key, "")).strip() for key in ALL_HEADERS}
        errors: list[tuple[str | None, str, str]] = []
        for field in REQUIRED_HEADERS:
            if not data[field]:
                errors.append((field, "REQUIRED", f"{field.replace('_', ' ').title()} is required."))
        for field, value in data.items():
            if value and (len(value) > 500 or not SAFE_TEXT.match(value)):
                errors.append((field, "INVALID_TEXT", "Contains unsafe or overlong text."))
        if data["barcode"] and len(data["barcode"]) > 80:
            errors.append(("barcode", "BARCODE_TOO_LONG", "Barcode exceeds 80 characters."))
        for money in ("purchase_cost", "selling_price", "mrp"):
            if not data[money]:
                continue
            value = self._decimal(data[money])
            if value is None or value < 0 or value > Decimal("9999999999.99"):
                errors.append((money, "INVALID_MONEY", "Use a non-negative money value with at most two decimals."))
            else:
                data[money] = format(value.quantize(Decimal("0.01")), "f")
        try:
            quantity = int(data["quantity"])
            if quantity <= 0 or quantity > 1_000_000:
                raise ValueError
            data["quantity"] = str(quantity)
        except (TypeError, ValueError):
            errors.append(("quantity", "INVALID_QUANTITY", "Quantity must be a positive whole number no greater than 1,000,000."))
        if data["mrp"] and self._decimal(data["mrp"]) is not None and self._decimal(data["selling_price"]) is not None and self._decimal(data["selling_price"]) > self._decimal(data["mrp"]):
            errors.append(("selling_price", "SELLING_PRICE_EXCEEDS_MRP", "Selling price cannot exceed MRP."))
        return data, errors

    @staticmethod
    def _decimal(value: str) -> Decimal | None:
        try:
            return Decimal(value)
        except (InvalidOperation, ValueError):
            return None

    def _finish_validation(self, batch: OpeningStockImport, current_user: User, request_id: str | None) -> None:
        rows = self.db.query(OpeningStockImportRow).filter_by(opening_stock_import_id=batch.id).all()
        for row in rows:
            if row.validation_status != "VALID":
                continue
            data = row.normalized_data
            product = self.db.query(Product).filter(func.lower(Product.sku) == data["sku"].casefold()).first()
            variant = self.db.query(ProductVariant).filter(ProductVariant.store_id == batch.store_id, func.lower(ProductVariant.barcode) == data["barcode"].casefold()).first()
            internal_sku_variant = self.db.query(ProductVariant).filter(ProductVariant.store_id == batch.store_id, func.lower(ProductVariant.internal_sku) == data["sku"].casefold()).first()
            if product and product.store_id != batch.store_id:
                row.validation_status = "INVALID"
                self._error(batch.id, row.id, row.row_number, "sku", "SKU_FOREIGN_STORE", "SKU already belongs to another store.")
            if variant and variant.product.sku and variant.product.sku.casefold() != data["sku"].casefold():
                row.validation_status = "INVALID"
                self._error(batch.id, row.id, row.row_number, "barcode", "BARCODE_PRODUCT_CONFLICT", "Barcode belongs to a different SKU.")
            if internal_sku_variant and (not variant or internal_sku_variant.id != variant.id):
                row.validation_status = "INVALID"
                self._error(batch.id, row.id, row.row_number, "sku", "VARIANT_SKU_CONFLICT", "SKU is already used by a different variant in this store.")
            if product and product.store_id == batch.store_id:
                expected = (data["category"].casefold(), data["subcategory"].casefold(), data["brand"].casefold())
                actual = (product.category.name.casefold(), product.subcategory.name.casefold(), product.brand.name.casefold())
                if actual != expected:
                    row.validation_status = "INVALID"
                    self._error(batch.id, row.id, row.row_number, "category", "CATALOG_CONFLICT", "Existing SKU has different category, subcategory, or brand values.")
        errors = self.db.query(OpeningStockImportError).filter_by(opening_stock_import_id=batch.id).all()
        backup_ok, backup_evidence = self._backup_gate()
        batch.backup_evidence = backup_evidence
        if not backup_ok:
            self._error(batch.id, None, None, None, "BACKUP_REQUIRED", "A current successful database backup must be recorded before this import can be posted.")
            errors.append(None)  # Counts as a validation blocker without exposing operational details.
        batch.row_count = len(rows)
        batch.error_count = len(errors)
        batch.valid_row_count = sum(1 for row in rows if row.validation_status == "VALID")
        batch.total_quantity = sum(int(row.normalized_data.get("quantity", "0") or 0) for row in rows if row.validation_status == "VALID")
        batch.total_cost_value = sum((Decimal(row.normalized_data["purchase_cost"]) * int(row.normalized_data["quantity"]) for row in rows if row.validation_status == "VALID"), Decimal("0"))
        batch.total_retail_value = sum((Decimal(row.normalized_data["selling_price"]) * int(row.normalized_data["quantity"]) for row in rows if row.validation_status == "VALID"), Decimal("0"))
        batch.validation_summary = {"required_headers": list(REQUIRED_HEADERS), "backup_gate_passed": backup_ok, "preview_generated_at": datetime.now(timezone.utc).isoformat()}
        batch.status = OpeningStockImportStatus.READY_TO_CONFIRM if batch.row_count and batch.error_count == 0 else OpeningStockImportStatus.REVIEW_REQUIRED
        self._audit(batch.id, "VALIDATED", current_user.id, request_id, {"row_count": batch.row_count, "error_count": batch.error_count, "backup_gate_passed": backup_ok})

    def _post_row(self, batch: OpeningStockImport, row: OpeningStockImportRow, current_user: User, request_id: str | None) -> Counter[str]:
        data = row.normalized_data
        store_id = batch.store_id
        result: Counter[str] = Counter()
        barcode = data["barcode"]
        variant = self.db.query(ProductVariant).filter(ProductVariant.store_id == store_id, func.lower(ProductVariant.barcode) == barcode.casefold()).with_for_update().first()
        product = variant.product if variant else self.db.query(Product).filter(Product.store_id == store_id, func.lower(Product.sku) == data["sku"].casefold()).with_for_update().first()
        if variant and product and product.sku and product.sku.casefold() != data["sku"].casefold():
            raise conflict(f"Row {row.row_number} barcode belongs to a different SKU.", "OPENING_STOCK_BARCODE_CONFLICT")
        category = self._catalog(Category, store_id, data["category"])
        if not category:
            category = Category(store_id=store_id, name=data["category"])
            self.db.add(category); self.db.flush()
        subcategory = self.db.query(SubCategory).filter(SubCategory.store_id == store_id, SubCategory.category_id == category.id, func.lower(SubCategory.name) == data["subcategory"].casefold()).first()
        if not subcategory:
            subcategory = SubCategory(store_id=store_id, category_id=category.id, name=data["subcategory"])
            self.db.add(subcategory); self.db.flush()
        brand = self.db.query(Brand).filter(Brand.store_id == store_id, Brand.category_id == category.id, func.lower(Brand.name) == data["brand"].casefold()).first()
        if not brand:
            brand = Brand(store_id=store_id, category_id=category.id, name=data["brand"])
            self.db.add(brand); self.db.flush()
        cost, selling = Decimal(data["purchase_cost"]), Decimal(data["selling_price"])
        mrp = Decimal(data["mrp"]) if data.get("mrp") else None
        if not product:
            product = Product(store_id=store_id, category_id=category.id, subcategory_id=subcategory.id, brand_id=brand.id, sku=data["sku"], name=data["product_name"], size=data.get("size") or None, color=data.get("color") or None, purchase_price=cost, selling_price=selling, pricing_type=PricingType.MRP if mrp is not None else PricingType.OWN_PRICE, mrp=mrp, barcode=barcode, hsn_code=data.get("hsn_code") or None, gst_rate=self._decimal(data.get("gst_rate", "")), description=data.get("description") or None, unit=data.get("unit") or "Each", warehouse=data.get("warehouse") or None)
            self.db.add(product); self.db.flush()
            result["created_products"] += 1
        if not variant:
            identity = "|".join((str(product.id), data.get("size", "").casefold(), data.get("color", "").casefold(), data.get("style_code", "").casefold(), data["sku"].casefold(), barcode.casefold(), str(mrp or selling), str(selling)))
            variant = ProductVariant(store_id=store_id, product_id=product.id, size=data.get("size") or None, color=data.get("color") or None, style_code=data.get("style_code") or None, internal_sku=data["sku"], barcode=barcode, identity_key=identity, mrp=mrp, selling_price=selling, last_purchase_cost=cost, average_cost=cost, current_stock=0)
            self.db.add(variant); self.db.flush()
            result["created_variants"] += 1
        mapping = self.db.query(ProductBarcode).filter(ProductBarcode.store_id == store_id, func.lower(ProductBarcode.barcode) == barcode.casefold()).with_for_update().first()
        if mapping and mapping.product_variant_id != variant.id:
            raise conflict(f"Row {row.row_number} barcode mapping changed during posting.", "OPENING_STOCK_BARCODE_MAPPING_CONFLICT")
        if not mapping:
            self.db.add(ProductBarcode(store_id=store_id, product_id=product.id, product_variant_id=variant.id, barcode=barcode, mrp=mrp, default_selling_price=selling, verified=True, verified_by=current_user.id, verified_at=datetime.now(timezone.utc)))
            result["created_barcodes"] += 1
        quantity = int(data["quantity"])
        before = variant.current_stock
        variant.current_stock += quantity
        variant.last_purchase_cost = cost
        variant.average_cost = ((variant.average_cost * before) + (cost * quantity)) / (before + quantity) if before else cost
        product.current_stock += quantity
        product.purchase_price = cost
        product.selling_price = selling
        inventory = self.db.query(ProductInventory).filter_by(product_id=product.id, store_id=store_id).with_for_update().first()
        if not inventory:
            inventory = ProductInventory(product_id=product.id, store_id=store_id, current_stock=0)
            self.db.add(inventory)
        inventory.current_stock += quantity
        lot = InventoryCostLot(store_id=store_id, product_variant_id=variant.id, received_quantity=quantity, remaining_quantity=quantity, unit_purchase_cost=cost, effective_unit_cost=cost, lot_reference=f"Opening stock import {batch.id}")
        self.db.add(lot); self.db.flush()
        result["created_cost_lots"] += 1
        movement = StockHistory(product_id=product.id, product_variant_id=variant.id, purchase_cost_lot_id=lot.id, unit_cost=cost, store_id=store_id, movement_type=StockMovementType.OPENING_STOCK, qty=quantity, before_stock=before, after_stock=variant.current_stock, reference=f"Opening stock import {batch.id}", request_id=request_id, created_by=current_user.id)
        self.db.add(movement); self.db.flush()
        result["created_movements"] += 1
        row.product_id, row.product_variant_id, row.cost_lot_id, row.stock_history_id = product.id, variant.id, lot.id, movement.id
        return result

    def _backup_gate(self) -> tuple[bool, dict[str, Any]]:
        if self.settings.allow_test_opening_stock_import_bypass and self.settings.app_env.lower() in {"test", "testing"}:
            return True, {"status": "test_bypass", "database_backup": "not_checked"}
        status_read = BackupStatusService(self.settings.backup_status_dir).status()
        database = next((item for item in status_read.components if item.component == "database"), None)
        ok = bool(status_read.configured and database and database.available and database.status.lower() == "success")
        return ok, {"configured": status_read.configured, "database_backup": database.status if database else "unknown"}

    def _write_evidence_file(self, stored_filename: str, content: bytes) -> None:
        directory = self.settings.opening_stock_import_dir
        directory.mkdir(parents=True, exist_ok=True)
        path = directory / stored_filename
        path.write_bytes(content)

    def _error(self, import_id: UUID, row_id: UUID | None, row_number: int | None, field: str | None, code: str, message: str) -> None:
        self.db.add(OpeningStockImportError(opening_stock_import_id=import_id, opening_stock_import_row_id=row_id, row_number=row_number, field=field, code=code, message=message))

    def _audit(self, import_id: UUID, action: str, user_id: UUID, request_id: str | None, metadata: dict[str, Any]) -> None:
        self.db.add(OpeningStockImportAudit(opening_stock_import_id=import_id, action=action, performed_by=user_id, request_id=request_id, metadata_json=metadata))

    def _catalog(self, model: type[Category], store_id: UUID, name: str) -> Category | None:
        return self.db.query(model).filter(model.store_id == store_id, func.lower(model.name) == name.casefold()).first()

    @staticmethod
    def _store_id(current_user: User) -> UUID:
        if not current_user.store_id:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Current user is not assigned to a store.")
        return current_user.store_id

    @staticmethod
    def _report(batch: OpeningStockImport, already_completed: bool = False) -> OpeningStockImportReport:
        result = batch.validation_summary.get("post_result", {}) if batch.validation_summary else {}
        return OpeningStockImportReport(import_id=batch.id, status=batch.status, created_products=int(result.get("created_products", 0)), created_variants=int(result.get("created_variants", 0)), created_barcodes=int(result.get("created_barcodes", 0)), created_cost_lots=int(result.get("created_cost_lots", 0)), created_movements=int(result.get("created_movements", 0)), total_quantity=batch.total_quantity, total_cost_value=batch.total_cost_value, total_retail_value=batch.total_retail_value, already_completed=already_completed)
