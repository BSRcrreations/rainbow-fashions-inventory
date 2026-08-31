from __future__ import annotations

import csv
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Optional, Union
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Request, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, require_manager_or_owner, require_owner
from app.core.exceptions import forbidden
from app.models.enums import UserRole
from app.database.session import get_db
from app.models.user import User
from app.schemas.product import (
    ProductBulkBrandUpdate,
    ProductBulkDeleteRequest,
    ProductBulkCategoryUpdate,
    ProductBulkIds,
    ProductBulkPurgeTestDataRequest,
    ProductBulkStockUpdate,
    ProductCodeResponse,
    ProductCreate,
    ProductImportSummary,
    ProductListResponse,
    ProductRead,
    ProductUpdateAuditRead,
    ProductUpdate,
)
from app.services.product_service import ProductService
from app.services.product_deletion_service import ProductDeletionService


router = APIRouter(prefix="/products", tags=["Products"])


@router.get("", response_model=Union[list[ProductRead], ProductListResponse])
def list_products(
    skip: int = 0,
    limit: int = 100,
    page: int = 1,
    page_size: int = 25,
    paginated: bool = False,
    search: Optional[str] = None,
    category_id: Optional[UUID] = None,
    brand_id: Optional[UUID] = None,
    is_active: Optional[bool] = None,
    stock_status: Optional[str] = None,
    min_price: Optional[Decimal] = None,
    max_price: Optional[Decimal] = None,
    created_from: Optional[date] = None,
    created_to: Optional[date] = None,
    sort_by: str = "name",
    sort_dir: str = "asc",
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    service = ProductService(db)
    if paginated:
        return service.list_paginated(
            page,
            page_size,
            search,
            category_id,
            brand_id,
            is_active,
            stock_status,
            min_price,
            max_price,
            created_from,
            created_to,
            sort_by,
            sort_dir,
        )
    return service.list(
        skip,
        limit,
        search,
        category_id,
        brand_id,
        is_active,
        stock_status,
        min_price,
        max_price,
        created_from,
        created_to,
        sort_by,
        sort_dir,
    )


@router.get("/generate-code", response_model=ProductCodeResponse)
def generate_product_code(kind: str = "sku", db: Session = Depends(get_db), _: User = Depends(require_manager_or_owner)):
    return ProductCodeResponse(value=ProductService(db).generate_code(kind))


@router.get("/export")
def export_products(
    format: str = "csv",
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    service = ProductService(db)
    if format == "xlsx":
        content = service.export_xlsx()
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=products.xlsx"},
        )
    content = service.export_csv()
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=products.csv"},
    )


@router.get("/import-template")
def download_import_template(_: User = Depends(get_current_user), db: Session = Depends(get_db)):
    content = ProductService(db).template_csv()
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=product-import-template.csv"},
    )


@router.post("/import", response_model=ProductImportSummary)
async def import_products(
    file: UploadFile = File(...),
    update_existing: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_or_owner),
):
    content = await file.read()
    rows: list[dict[str, str]]
    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            from app.core.exceptions import bad_request

            raise bad_request("XLSX import requires openpyxl to be installed") from exc
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            rows = []
        else:
            headers = [str(value or "").strip() for value in values[0]]
            rows = [dict(zip(headers, [str(value or "") for value in row])) for row in values[1:]]
    else:
        text = content.decode("utf-8-sig")
        rows = list(csv.DictReader(StringIO(text)))
    return ProductService(db).import_products(rows, update_existing, current_user.store_id)


@router.post("", response_model=ProductRead, status_code=status.HTTP_201_CREATED)
def create_product(payload: ProductCreate, db: Session = Depends(get_db), current_user: User = Depends(require_manager_or_owner)):
    if payload.is_test_data and current_user.role != UserRole.OWNER:
        raise forbidden("Only an owner can mark a product as test data")
    return ProductService(db).create(payload, current_user.store_id)


@router.get("/barcode/{barcode}", response_model=ProductRead)
def get_product_by_barcode(barcode: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return ProductService(db).get_by_barcode(barcode, current_user.store_id)


@router.get("/{product_id}", response_model=ProductRead)
def get_product(product_id: UUID, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return ProductService(db).get(product_id)


@router.put("/{product_id}", response_model=ProductRead)
@router.patch("/{product_id}", response_model=ProductRead)
def update_product(product_id: UUID, payload: ProductUpdate, request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_manager_or_owner)):
    if payload.is_test_data is True and current_user.role != UserRole.OWNER:
        raise forbidden("Only an owner can mark a product as test data")
    return ProductService(db).update(product_id, payload, current_user.store_id, current_user, request.state.request_id)


@router.get("/{product_id}/audit", response_model=list[ProductUpdateAuditRead])
def list_product_update_audits(product_id: UUID, db: Session = Depends(get_db), current_user: User = Depends(require_manager_or_owner)):
    return ProductService(db).list_update_audits(product_id, current_user)


@router.post("/{product_id}/archive", response_model=ProductRead)
def archive_product(product_id: UUID, request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_manager_or_owner)):
    return ProductService(db).archive(product_id, current_user, request.state.request_id)


@router.post("/{product_id}/restore", response_model=ProductRead)
def restore_product(product_id: UUID, request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_manager_or_owner)):
    return ProductService(db).restore(product_id, current_user, request.state.request_id)


@router.get("/{product_id}/deletion-check")
def product_deletion_check(product_id: UUID, request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_owner)):
    return ProductDeletionService(db).check([product_id], current_user, request.state.request_id)


@router.post("/{product_id}/image", response_model=ProductRead)
async def upload_product_image(
    product_id: UUID,
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_or_owner),
):
    return await ProductService(db).upload_image(product_id, file, current_user.id, current_user, request.state.request_id)


@router.delete("/{product_id}/image", response_model=ProductRead)
def delete_product_image(product_id: UUID, request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_manager_or_owner)):
    return ProductService(db).delete_image(product_id, current_user, request.state.request_id)


@router.post("/bulk-delete-check")
def check_bulk_product_delete(
    payload: ProductBulkIds,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_owner),
):
    return ProductDeletionService(db).check(payload.product_ids, current_user, request.state.request_id)


@router.post("/bulk-delete")
def permanently_delete_products(
    payload: ProductBulkDeleteRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_owner),
):
    return ProductDeletionService(db).permanently_delete(
        payload.product_ids, payload.confirmation, payload.delete_password,
        request.headers.get("Idempotency-Key", ""), current_user, request.state.request_id,
        request.client.host if request.client else None,
    )


@router.post("/bulk-purge-test-data")
def purge_bulk_test_products(
    payload: ProductBulkPurgeTestDataRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_owner),
):
    return ProductDeletionService(db).purge_test_data(payload.product_ids, payload.confirmation, payload.reason, current_user, request.state.request_id)


@router.post("/bulk/delete", deprecated=True)
def bulk_delete_products_legacy(
    payload: ProductBulkDeleteRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_owner),
):
    """Compatibility alias. It now enforces the same typed owner confirmation."""
    return ProductDeletionService(db).permanently_delete(
        payload.product_ids, payload.confirmation, payload.delete_password,
        request.headers.get("Idempotency-Key", ""), current_user, request.state.request_id,
        request.client.host if request.client else None,
    )


@router.post("/bulk/category")
def bulk_update_product_category(payload: ProductBulkCategoryUpdate, db: Session = Depends(get_db), _: User = Depends(require_manager_or_owner)):
    return ProductService(db).bulk_update_category(payload)


@router.post("/bulk/brand")
def bulk_update_product_brand(payload: ProductBulkBrandUpdate, db: Session = Depends(get_db), _: User = Depends(require_manager_or_owner)):
    return ProductService(db).bulk_update_brand(payload)


@router.post("/bulk/stock")
def bulk_update_product_stock(payload: ProductBulkStockUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_manager_or_owner)):
    return ProductService(db).bulk_stock_update(payload, current_user)


@router.post("/bulk/export")
def bulk_export_products(payload: ProductBulkIds, format: str = "csv", db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    service = ProductService(db)
    if format == "xlsx":
        content = service.export_xlsx(payload.product_ids)
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=selected-products.xlsx"},
        )
    content = service.export_csv(payload.product_ids)
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=selected-products.csv"},
    )


@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT, deprecated=True)
def delete_product(product_id: UUID, db: Session = Depends(get_db), _: User = Depends(require_owner)) -> Response:
    raise HTTPException(status_code=status.HTTP_410_GONE, detail={"message": "Use the typed permanent-delete confirmation workflow.", "code": "PRODUCT_DELETE_CONFIRMATION_REQUIRED"})
